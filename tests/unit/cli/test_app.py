"""Tests for cli/app.py — Typer commands via CliRunner."""
from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from typer.testing import CliRunner

from doors_excel.cli.app import app

runner = CliRunner()


def _write_valid_config(tmp_path: Path) -> Path:
    data = {
        "modules": [
            {
                "module_path": "/proj/mod",
                "column_mappings": [
                    {"column": "Object Text", "attribute": "Object Text",
                     "attribute_type": "Text"},
                    {"column": "Short Name", "attribute": "Short Name",
                     "attribute_type": "String"},
                ],
            }
        ]
    }
    p = tmp_path / "config.json"
    p.write_text(json.dumps(data), encoding="utf-8")
    return p


def _write_valid_xlsx(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mod"
    ws.append(["Absolute Number", "Object Text", "Short Name"])
    ws.append([10, "Hello", "hi"])
    p = tmp_path / "data.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# validate — config only
# ---------------------------------------------------------------------------

class TestValidateConfigOnly:
    def test_valid_config_exits_zero(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        result = runner.invoke(app, ["validate", "--config", str(cfg)])
        assert result.exit_code == 0

    def test_missing_config_exits_one(self, tmp_path: Path) -> None:
        result = runner.invoke(app, ["validate", "--config", str(tmp_path / "missing.json")])
        assert result.exit_code == 1

    def test_bad_json_config_exits_one(self, tmp_path: Path) -> None:
        p = tmp_path / "bad.json"
        p.write_text("{broken", encoding="utf-8")
        result = runner.invoke(app, ["validate", "--config", str(p)])
        assert result.exit_code == 1

    def test_no_args_exits_nonzero(self) -> None:
        result = runner.invoke(app, ["validate"])
        assert result.exit_code != 0

    def test_quiet_suppresses_ok_message(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        result = runner.invoke(app, ["validate", "--config", str(cfg), "--quiet"])
        assert result.exit_code == 0
        assert "Config OK" not in result.output


# ---------------------------------------------------------------------------
# validate — config + file
# ---------------------------------------------------------------------------

class TestValidateWithFile:
    def test_valid_xlsx_exits_zero(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        xlsx = _write_valid_xlsx(tmp_path)
        result = runner.invoke(app, ["validate", "--config", str(cfg), "--file", str(xlsx)])
        assert result.exit_code == 0

    def test_file_without_config_exits_one(self, tmp_path: Path) -> None:
        xlsx = _write_valid_xlsx(tmp_path)
        result = runner.invoke(app, ["validate", "--file", str(xlsx)])
        assert result.exit_code == 1

    def test_xlsx_with_invalid_string_exits_one(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "mod"
        ws.append(["Absolute Number", "Object Text", "Short Name"])
        ws.append([10, "ok", "x" * 1025])  # too long for String
        p = tmp_path / "bad.xlsx"
        wb.save(p)
        result = runner.invoke(app, ["validate", "--config", str(cfg), "--file", str(p)])
        assert result.exit_code == 1


# ---------------------------------------------------------------------------
# Stub commands (gui is still a stub)
# ---------------------------------------------------------------------------

class TestStubCommands:
    def test_gui_command_exists(self) -> None:
        result = runner.invoke(app, ["gui", "--help"])
        assert result.exit_code == 0


# ---------------------------------------------------------------------------
# export (real implementation)
# ---------------------------------------------------------------------------

class TestExportCommand:
    def test_export_succeeds_with_mocked_api(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.export_module_api", return_value=out):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(app, ["export", "--config", str(cfg), "--output", str(out)])
        assert result.exit_code == 0

    def test_export_missing_config_exits_one(self, tmp_path: Path) -> None:
        result = runner.invoke(
            app,
            ["export", "--config", str(tmp_path / "missing.json"), "--output", str(tmp_path / "out.xlsx")],
        )
        assert result.exit_code == 1

    def test_export_doors_unavailable_exits_one(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn:
            MockConn.open.side_effect = Exception("COM unavailable")
            result = runner.invoke(app, ["export", "--config", str(cfg), "--output", str(out)])
        assert result.exit_code == 1

    def test_export_quiet_suppresses_success_message(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.export_module_api", return_value=out):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["export", "--config", str(cfg), "--output", str(out), "--quiet"],
            )
        assert result.exit_code == 0
        assert "Exported" not in result.output

    def test_export_creates_session_db_and_json(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        out.touch()  # export_module_api needs file to exist for SessionManager.create

        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.export_module_api", return_value=out) as mock_export:
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(app, ["export", "--config", str(cfg), "--output", str(out)])

        assert result.exit_code == 0
        # Check that session_manager was passed to the API
        call_kwargs = mock_export.call_args.kwargs
        assert "session_manager" in call_kwargs
        assert call_kwargs["session_manager"] is not None

    def test_export_starts_keepalive_watchdog(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        out.touch()
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.export_module_api", return_value=out), \
             patch("doors_excel.cli.app.KeepAliveWatchdog") as MockWatchdog:
            MockConn.open.return_value = MagicMock()
            mock_wd = MagicMock()
            MockWatchdog.return_value = mock_wd
            result = runner.invoke(app, ["export", "--config", str(cfg), "--output", str(out)])
        assert result.exit_code == 0
        MockWatchdog.assert_called_once()
        mock_wd.start.assert_called_once()
        mock_wd.stop.assert_called_once()

    def test_export_watchdog_stopped_on_doors_error(self, tmp_path: Path) -> None:
        from doors_excel.common.exceptions import DoorsExcelError

        cfg = _write_valid_config(tmp_path)
        out = tmp_path / "out.xlsx"
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.export_module_api",
                   side_effect=DoorsExcelError("boom")), \
             patch("doors_excel.cli.app.KeepAliveWatchdog") as MockWatchdog:
            MockConn.open.return_value = MagicMock()
            mock_wd = MagicMock()
            MockWatchdog.return_value = mock_wd
            result = runner.invoke(
                app,
                ["export", "--config", str(cfg), "--output", str(out)],
            )
        assert result.exit_code == 1
        mock_wd.stop.assert_called_once()


# ---------------------------------------------------------------------------
# rollback (real implementation)
# ---------------------------------------------------------------------------

class TestRollbackCommand:
    def test_rollback_succeeds_with_mocked_api(self, tmp_path: Path) -> None:
        db_path = str(tmp_path / "test.db")
        session_file = tmp_path / ".doors_session.json"
        session_file.write_text(
            json.dumps({"session_id": "s1", "db_path": db_path}), encoding="utf-8"
        )
        out = tmp_path / "rollback.xlsx"
        with patch("doors_excel.cli.app.generate_rollback_excel_api", return_value=out):
            result = runner.invoke(
                app,
                ["rollback", "--session", str(session_file), "--output", str(out)],
            )
        assert result.exit_code == 0

    def test_rollback_missing_session_file_exits_one(self, tmp_path: Path) -> None:
        result = runner.invoke(
            app,
            ["rollback", "--session", str(tmp_path / "no.json")],
        )
        assert result.exit_code == 1


# ---------------------------------------------------------------------------
# import (real implementation)
# ---------------------------------------------------------------------------

class TestImportCommand:
    def _write_import_xlsx(self, tmp_path: Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "mod"
        ws.append(["Absolute Number", "Object Text", "Short Name"])
        ws.append([1, "updated text", "sn"])
        p = tmp_path / "data.xlsx"
        wb.save(p)
        return p

    def test_import_succeeds_with_mocked_api(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg)],
            )
        assert result.exit_code == 0

    def test_import_exits_one_when_conflicts_present_and_doors_wins_policy(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=0, conflict_count=2, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=0):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--policy", "doors-wins"],
            )
        assert result.exit_code == 1

    def test_import_missing_config_exits_one(self, tmp_path: Path) -> None:
        xlsx = self._write_import_xlsx(tmp_path)
        result = runner.invoke(
            app,
            ["import", "--file", str(xlsx), "--config", str(tmp_path / "missing.json")],
        )
        assert result.exit_code == 1

    def test_import_quiet_suppresses_success_message(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--quiet"],
            )
        assert result.exit_code == 0
        assert "Applied" not in result.output

    def test_purge_without_force_exits_one(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=1, updated_count=0, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=0):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg),
                 "--deletion-policy", "purge"],
            )
        assert result.exit_code == 1
        assert "force" in result.output.lower()

    def test_import_starts_keepalive_watchdog(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1), \
             patch("doors_excel.cli.app.KeepAliveWatchdog") as MockWatchdog:
            MockConn.open.return_value = MagicMock()
            mock_wd = MagicMock()
            MockWatchdog.return_value = mock_wd
            result = runner.invoke(app, ["import", "--file", str(xlsx), "--config", str(cfg)])
        assert result.exit_code == 0
        MockWatchdog.assert_called_once()
        mock_wd.start.assert_called_once()
        mock_wd.stop.assert_called_once()

    def test_import_watchdog_stopped_on_doors_error(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary
        from doors_excel.common.exceptions import DoorsExcelError

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api",
                   side_effect=DoorsExcelError("fail")), \
             patch("doors_excel.cli.app.KeepAliveWatchdog") as MockWatchdog:
            MockConn.open.return_value = MagicMock()
            mock_wd = MagicMock()
            MockWatchdog.return_value = mock_wd
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg)],
            )
        assert result.exit_code == 1
        mock_wd.stop.assert_called_once()

    def test_import_ole_flag_passed_to_execute(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1) as mock_exec, \
             patch("doors_excel.cli.app.KeepAliveWatchdog"):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--accept-ole-overwrites"],
            )
        assert result.exit_code == 0
        call_kwargs = mock_exec.call_args.kwargs
        assert call_kwargs.get("accept_ole_overwrites") is True

    def test_purge_with_children_prints_warning(self, tmp_path: Path) -> None:
        """When purging objects that have children in DOORS, user sees a warning count."""
        from doors_excel.core.diff.summary import DiffSummary

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        stats = DiffSummary(new_count=0, deleted_count=3, updated_count=0, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid1", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=3), \
             patch("doors_excel.cli.app.KeepAliveWatchdog"), \
             patch("doors_excel.cli.app._count_staged_children", return_value=12):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg),
                 "--deletion-policy", "purge", "--force"],
            )
        assert result.exit_code == 0
        assert "12" in result.output
        assert "cascade" in result.output.lower() or "children" in result.output.lower()


# ---------------------------------------------------------------------------
# import — session recovery
# ---------------------------------------------------------------------------

class TestImportSessionRecovery:
    def _write_import_xlsx(self, tmp_path: Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "mod"
        ws.append(["Absolute Number", "Object Text", "Short Name"])
        ws.append([1, "text", "sn"])
        p = tmp_path / "data.xlsx"
        wb.save(p)
        return p

    def test_existing_session_without_flag_exits_one(self, tmp_path: Path) -> None:
        """Leftover session file aborts unless --resume or --discard-session given."""
        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        # write a fake session sidecar next to the xlsx
        session_file = xlsx.parent / ".session.json"
        session_file.write_text('{"session_id": "old", "db_path": "x.db"}', encoding="utf-8")
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn:
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg)],
            )
        assert result.exit_code == 1
        assert "resume" in result.output.lower() or "session" in result.output.lower()

    def test_discard_session_flag_removes_file_and_proceeds(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary
        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        session_file = xlsx.parent / ".session.json"
        session_file.write_text('{"session_id": "old", "db_path": "x.db"}', encoding="utf-8")

        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid2", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1), \
             patch("doors_excel.cli.app.KeepAliveWatchdog"):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--discard-session"],
            )
        assert result.exit_code == 0
        assert not session_file.exists()

    def test_resume_validates_session_and_exits_zero(self, tmp_path: Path) -> None:
        """--resume validates the session via SessionManager.resume() and exits 0."""
        from doors_excel.core.diff.summary import DiffSummary
        from doors_excel.api.sessions import SessionInfo

        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        session_file = xlsx.parent / ".session.json"
        session_file.write_text('{"session_id": "old", "db_path": "x.db"}', encoding="utf-8")

        fake_info = SessionInfo(
            session_id="old",
            excel_path=xlsx,
            doors_module="/proj/mod",
            excel_sha256="abc",
            module_version="current",
            db_path=tmp_path / "x.db",
        )
        fake_stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)

        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app._SessionMgr") as MockMgr, \
             patch("doors_excel.cli.app._run_diff_api", return_value=fake_stats):
            MockConn.open.return_value = MagicMock()
            mock_mgr_instance = MagicMock()
            MockMgr.return_value = mock_mgr_instance
            mock_mgr_instance.resume.return_value = fake_info
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--resume"],
            )
        assert result.exit_code == 0
        mock_mgr_instance.resume.assert_called_once()
        MockConn.open.assert_not_called()

    def test_resume_and_discard_session_together_exits_one(self, tmp_path: Path) -> None:
        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        result = runner.invoke(
            app,
            ["import", "--file", str(xlsx), "--config", str(cfg),
             "--resume", "--discard-session"],
        )
        assert result.exit_code == 1

    def test_discard_session_removes_db_too(self, tmp_path: Path) -> None:
        from doors_excel.core.diff.summary import DiffSummary
        cfg = _write_valid_config(tmp_path)
        xlsx = self._write_import_xlsx(tmp_path)
        session_file = xlsx.parent / ".session.json"
        session_file.write_text('{"session_id": "old", "db_path": "x.db"}', encoding="utf-8")
        db_file = xlsx.parent / (xlsx.stem + ".db")
        db_file.write_text("fake db", encoding="utf-8")

        stats = DiffSummary(new_count=0, deleted_count=0, updated_count=1, conflict_count=0, moved_count=0, baseline_mismatch_count=0)
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.stage_import_api", return_value=("sid2", stats)), \
             patch("doors_excel.cli.app.execute_import_api", return_value=1), \
             patch("doors_excel.cli.app.KeepAliveWatchdog"):
            MockConn.open.return_value = MagicMock()
            result = runner.invoke(
                app,
                ["import", "--file", str(xlsx), "--config", str(cfg), "--discard-session"],
            )
        assert result.exit_code == 0
        assert not session_file.exists()
        # The old fake-content DB was deleted; the import created a new valid SQLite DB in its place
        assert db_file.exists()
        assert db_file.read_text(encoding="utf-8", errors="replace") != "fake db"
