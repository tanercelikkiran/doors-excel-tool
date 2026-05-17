"""Unit tests for api/export.py — DoorsExporter is fully mocked."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from doors_excel.core.validation.models import ColumnMapping, ModuleConfig
from doors_excel.api.sessions import SessionManager


def _make_module_config(attrs: list[str] | None = None) -> ModuleConfig:
    mappings = [
        ColumnMapping(column=a, attribute=a, attribute_type="Text")
        for a in (attrs or ["Object Text"])
    ]
    return ModuleConfig(module_path="/proj/mod", column_mappings=mappings)


def _raw_rows(object_id: int = 1, level: int = 1, attrs: list[str] | None = None) -> list[dict]:
    """Return fake DoorsExporter output rows for one object."""
    attrs = attrs or ["Object Text"]
    return [
        {
            "object_id": object_id,
            "level": level,
            "parent_id": None,
            "has_ole": 0,
            "object_type": "OBJECT",
            "attribute": a,
            "value": f"Value of {a}",
            "rtf_value": "",
            "md_hash": None,
            "parent_absno": None,
            "row_position": None,
            "col_position": None,
        }
        for a in attrs
        # rtf_value="" skips RTF→Markdown conversion for most tests;
        # tests that need conversion override rtf_value explicitly.
    ]


class TestExportModule:
    def test_returns_path_to_xlsx(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_excel_has_header_row_with_object_id_column(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "Absolute Number" in headers

    def test_excel_contains_attribute_column(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        cfg = _make_module_config(["Object Text", "Short Name"])
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            out = export_module("/proj/mod", cfg, tmp_path / "out.xlsx", doors_conn=mock_conn)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "Object Text" in headers
        assert "Short Name" in headers

    def test_one_object_produces_one_data_row(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(object_id=42),
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        assert 42 in data_rows[0]

    def test_rtf_converted_to_markdown_for_text_attributes(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        rows = _raw_rows()
        rows[0]["rtf_value"] = "{\\rtf1\\ansi {\\b Bold text}}"
        rows[0]["value"] = "Bold text"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=rows,
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(["Object Text"]),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        col_idx = headers.index("Object Text") + 1
        cell_val = ws.cell(row=2, column=col_idx).value
        assert cell_val is not None
        assert "{\\rtf1" not in str(cell_val)

    def test_two_objects_produce_two_data_rows(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        rows = _raw_rows(object_id=1) + _raw_rows(object_id=2)
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=rows,
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=[],
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "sub" / "out.xlsx",
                doors_conn=mock_conn,
            )
        assert out.exists()

    def test_module_path_embedded_in_workbook_metadata(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module
        from doors_excel.infrastructure.excel.metadata import get_module_path

        mock_conn = object()
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        module_path = get_module_path(wb, ws)
        assert module_path == "/proj/mod"

    def test_excel_has_object_type_column(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            out = export_module(
                "/proj/mod", _make_module_config(), tmp_path / "out.xlsx",
                doors_conn=object(),
            )
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
        assert "Object Type" in headers

    def test_regular_object_rows_have_object_marker(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(object_id=5),
        ):
            out = export_module(
                "/proj/mod", _make_module_config(), tmp_path / "out.xlsx",
                doors_conn=object(),
            )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ot_col_idx = headers.index("Object Type")
        data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert data_row[ot_col_idx] == "OBJECT"

    def test_excel_has_doors_position_system_columns(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            out = export_module(
                "/proj/mod", _make_module_config(), tmp_path / "out.xlsx",
                doors_conn=object(),
            )
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
        assert "_DOORS_Parent_AbsNo" in headers
        assert "_DOORS_Row_Position" in headers
        assert "_DOORS_Col_Position" in headers


# ---------------------------------------------------------------------------
# Session-aware export
# ---------------------------------------------------------------------------

class TestSessionAwareExport:
    def test_staging_doors_populated_when_session_manager_provided(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        db_path = tmp_path / "test.db"
        mock_conn = object()
        rows = _raw_rows(object_id=7, attrs=["Object Text"])
        with patch("doors_excel.api.export.DoorsExporter.export_module", return_value=rows):
            mgr = SessionManager(db_path)
            export_module(
                "/proj/mod",
                _make_module_config(["Object Text"]),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
                session_manager=mgr,
            )
            staging = mgr.conn.execute(
                "SELECT * FROM staging_doors WHERE attribute = 'Object Text'"
            ).fetchall()
        assert len(staging) == 1
        assert staging[0]["object_id"] == 7

    def test_staging_baseline_populated_when_session_manager_provided(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        db_path = tmp_path / "test.db"
        mock_conn = object()
        rows = _raw_rows(object_id=7, attrs=["Object Text"])
        with patch("doors_excel.api.export.DoorsExporter.export_module", return_value=rows):
            mgr = SessionManager(db_path)
            export_module(
                "/proj/mod",
                _make_module_config(["Object Text"]),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
                session_manager=mgr,
            )
            baseline = mgr.conn.execute(
                "SELECT * FROM staging_baseline WHERE attribute = 'Object Text'"
            ).fetchall()
        assert len(baseline) == 1

    def test_rollback_snapshots_populated_when_session_manager_provided(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        db_path = tmp_path / "test.db"
        mock_conn = object()
        rows = _raw_rows(object_id=7, attrs=["Object Text"])
        with patch("doors_excel.api.export.DoorsExporter.export_module", return_value=rows):
            mgr = SessionManager(db_path)
            export_module(
                "/proj/mod",
                _make_module_config(["Object Text"]),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
                session_manager=mgr,
            )
            snaps = mgr.conn.execute("SELECT * FROM rollback_snapshots").fetchall()
        assert len(snaps) == 1
        assert snaps[0]["object_id"] == 7

    def test_no_session_created_when_session_manager_not_provided(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        mock_conn = object()
        with patch("doors_excel.api.export.DoorsExporter.export_module", return_value=_raw_rows()):
            out = export_module(
                "/proj/mod",
                _make_module_config(),
                tmp_path / "out.xlsx",
                doors_conn=mock_conn,
            )
        assert out.exists()  # unchanged behaviour


# ---------------------------------------------------------------------------
# Sheet protection
# ---------------------------------------------------------------------------

class TestSheetProtection:
    def _make_mod_cfg(self, read_only_cols: list[str] | None = None) -> ModuleConfig:
        return ModuleConfig(
            module_path="/proj/Mod",
            column_mappings=[
                ColumnMapping(
                    column="Object Text",
                    attribute="Object Text",
                    attribute_type="Text",
                    read_only=("Object Text" in (read_only_cols or [])),
                ),
                ColumnMapping(
                    column="Short Name",
                    attribute="Short Name",
                    attribute_type="String",
                    read_only=("Short Name" in (read_only_cols or [])),
                ),
            ],
        )

    def test_protection_not_applied_by_default(self, tmp_path):
        """No sheet_protection param → sheet.protection.sheet is False."""
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            export_module(
                "/proj/Mod", self._make_mod_cfg(), out,
                doors_conn=object(),
            )

        wb = openpyxl.load_workbook(out)
        assert not wb.active.protection.sheet

    def test_protection_applied_when_enabled(self, tmp_path):
        """sheet_protection=True → sheet.protection.sheet is True."""
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            export_module(
                "/proj/Mod", self._make_mod_cfg(), out,
                doors_conn=object(),
                sheet_protection=True,
            )

        wb = openpyxl.load_workbook(out)
        assert wb.active.protection.sheet

    def test_read_only_column_cells_locked(self, tmp_path):
        """Columns with read_only=True → data cells are locked."""
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            export_module(
                "/proj/Mod",
                self._make_mod_cfg(read_only_cols=["Object Text", "Short Name"]),
                out,
                doors_conn=object(),
                sheet_protection=True,
            )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ot_col = headers.index("Object Text") + 1
        # data row 2, Object Text is read_only → locked=True (or not explicitly False)
        assert ws.cell(2, ot_col).protection.locked is not False

    def test_editable_column_cells_unlocked(self, tmp_path):
        """Columns with read_only=False → data cells are unlocked."""
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            export_module(
                "/proj/Mod",
                self._make_mod_cfg(read_only_cols=[]),  # all editable
                out,
                doors_conn=object(),
                sheet_protection=True,
            )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ot_col = headers.index("Object Text") + 1
        # Object Text is read_only=False → unlocked
        assert ws.cell(2, ot_col).protection.locked is False

    def test_mixed_columns_locked_and_unlocked(self, tmp_path):
        """One read_only=True column stays locked, one read_only=False is unlocked — in same workbook."""
        from doors_excel.api.export import export_module

        mixed_cfg = ModuleConfig(
            module_path="/proj/Mod",
            column_mappings=[
                ColumnMapping(column="Object Text", attribute="Object Text", attribute_type="Text", read_only=False),
                ColumnMapping(column="Short Name", attribute="Short Name", attribute_type="String", read_only=True),
            ],
        )
        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(attrs=["Object Text", "Short Name"]),
        ):
            export_module(
                "/proj/Mod", mixed_cfg, out,
                doors_conn=object(),
                sheet_protection=True,
            )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ot_col = headers.index("Object Text") + 1
        sn_col = headers.index("Short Name") + 1
        # Object Text is editable → explicitly unlocked
        assert ws.cell(2, ot_col).protection.locked is False
        # Short Name is read_only → locked (True or None = openpyxl default locked)
        assert ws.cell(2, sn_col).protection.locked is not False


# ---------------------------------------------------------------------------
# Source baseline column
# ---------------------------------------------------------------------------

class TestSourceBaselineColumn:
    def _make_mod_cfg(self) -> ModuleConfig:
        return ModuleConfig(
            module_path="/proj/Mod",
            column_mappings=[
                ColumnMapping(column="Object Text", attribute="Object Text", attribute_type="Text"),
            ],
        )

    def test_source_baseline_column_absent_by_default(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            export_module(
                "/proj/Mod",
                self._make_mod_cfg(),
                out,
                doors_conn=object(),
            )

        wb = openpyxl.load_workbook(out)
        headers = [wb.active.cell(1, c).value for c in range(1, wb.active.max_column + 1)]
        assert "Source Baseline" not in headers

    def test_source_baseline_column_present_when_enabled(self, tmp_path: Path) -> None:
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            export_module(
                "/proj/Mod",
                self._make_mod_cfg(),
                out,
                doors_conn=object(),
                include_source_baseline=True,
                baseline="v1.0",
            )

        wb = openpyxl.load_workbook(out)
        headers = [wb.active.cell(1, c).value for c in range(1, wb.active.max_column + 1)]
        assert "Source Baseline" in headers
        sb_col = headers.index("Source Baseline") + 1
        assert wb.active.cell(2, sb_col).value == "v1.0"

    def test_source_baseline_default_baseline_value(self, tmp_path: Path) -> None:
        """When no baseline= is passed, the column value should be 'current'."""
        from doors_excel.api.export import export_module

        out = tmp_path / "out.xlsx"
        with patch(
            "doors_excel.api.export.DoorsExporter.export_module",
            return_value=_raw_rows(),
        ):
            export_module(
                "/proj/Mod",
                self._make_mod_cfg(),
                out,
                doors_conn=object(),
                include_source_baseline=True,
            )

        wb = openpyxl.load_workbook(out)
        headers = [wb.active.cell(1, c).value for c in range(1, wb.active.max_column + 1)]
        sb_col = headers.index("Source Baseline") + 1
        assert wb.active.cell(2, sb_col).value == "current"
