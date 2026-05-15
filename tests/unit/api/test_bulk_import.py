"""Tests for bulk import worksheet-module resolution and staging (REQ-FUN-215)."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from doors_excel.core.validation.models import ColumnMapping, ModuleConfig, ProjectConfig


def _make_project_cfg(module_paths: list[str]) -> ProjectConfig:
    return ProjectConfig(
        modules=[
            ModuleConfig(
                module_path=mp,
                column_mappings=[
                    ColumnMapping(column="Object Text", attribute="Object Text", attribute_type="Text"),
                ],
            )
            for mp in module_paths
        ]
    )


def _make_xlsx_with_sheets(tmp_path: Path, sheet_configs: list[tuple[str, str | None]]) -> Path:
    """Create xlsx with named sheets. sheet_configs = [(sheet_name, module_path_or_None)]."""
    from doors_excel.infrastructure.excel.metadata import set_module_path

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, module_path in sheet_configs:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Absolute Number", "Object Text"])
        ws.append([1, "hello"])
        if module_path is not None:
            set_module_path(wb, ws, module_path)
    p = tmp_path / "test.xlsx"
    wb.save(p)
    return p


class TestResolveWorksheetModule:
    def test_match_by_custom_property(self, tmp_path: Path) -> None:
        """Custom property _DOORS_ModulePath takes priority."""
        from doors_excel.api.import_ import resolve_worksheet_module

        project_cfg = _make_project_cfg(["/proj/ModA"])
        xlsx = _make_xlsx_with_sheets(tmp_path, [("SomeSheet", "/proj/ModA")])

        wb = openpyxl.load_workbook(xlsx)
        ws = wb["SomeSheet"]
        result = resolve_worksheet_module(wb, ws, project_cfg)
        assert result is not None
        assert result.module_path == "/proj/ModA"

    def test_match_by_sheet_name(self, tmp_path: Path) -> None:
        """Falls back to make_sheet_name match when no custom property."""
        from doors_excel.api.import_ import resolve_worksheet_module
        from doors_excel.infrastructure.excel.naming import make_sheet_name

        project_cfg = _make_project_cfg(["/proj/ModA"])
        sheet_name = make_sheet_name("ModA", "/proj/ModA")
        xlsx = _make_xlsx_with_sheets(tmp_path, [(sheet_name, None)])

        wb = openpyxl.load_workbook(xlsx)
        ws = wb[sheet_name]
        result = resolve_worksheet_module(wb, ws, project_cfg)
        assert result is not None
        assert result.module_path == "/proj/ModA"

    def test_no_match_returns_none(self, tmp_path: Path) -> None:
        from doors_excel.api.import_ import resolve_worksheet_module

        project_cfg = _make_project_cfg(["/proj/ModA"])
        xlsx = _make_xlsx_with_sheets(tmp_path, [("UnrelatedSheet", None)])

        wb = openpyxl.load_workbook(xlsx)
        ws = wb["UnrelatedSheet"]
        result = resolve_worksheet_module(wb, ws, project_cfg)
        assert result is None

    def test_match_by_case_insensitive_name(self, tmp_path: Path) -> None:
        """Priority 3: case-insensitive module name match against sheet title."""
        from doors_excel.api.import_ import resolve_worksheet_module

        project_cfg = _make_project_cfg(["/proj/ModA"])
        # Sheet title is lowercase "moda" — should still match module "ModA"
        xlsx = _make_xlsx_with_sheets(tmp_path, [("moda", None)])

        wb = openpyxl.load_workbook(xlsx)
        ws = wb["moda"]
        result = resolve_worksheet_module(wb, ws, project_cfg)
        assert result is not None
        assert result.module_path == "/proj/ModA"

    def test_custom_property_wins_over_sheet_name(self, tmp_path: Path) -> None:
        """Priority 1 (custom property) takes precedence over priority 2 (sheet name)."""
        from doors_excel.api.import_ import resolve_worksheet_module
        from doors_excel.infrastructure.excel.naming import make_sheet_name

        # Sheet title matches ModB's make_sheet_name but has custom property for ModA
        project_cfg = _make_project_cfg(["/proj/ModA", "/proj/ModB"])
        sheet_b_name = make_sheet_name("ModB", "/proj/ModB")
        # This sheet looks like ModB by name but is actually ModA by custom property
        xlsx = _make_xlsx_with_sheets(tmp_path, [(sheet_b_name, "/proj/ModA")])

        wb = openpyxl.load_workbook(xlsx)
        ws = wb[sheet_b_name]
        result = resolve_worksheet_module(wb, ws, project_cfg)
        assert result is not None
        assert result.module_path == "/proj/ModA"  # custom property wins


class TestBulkStageImports:
    def test_bulk_stage_imports_returns_one_result_per_matched_sheet(self, tmp_path: Path) -> None:
        from doors_excel.api.import_ import bulk_stage_imports
        from doors_excel.infrastructure.excel.naming import make_sheet_name

        sheet_a = make_sheet_name("ModA", "/proj/ModA")
        sheet_b = make_sheet_name("ModB", "/proj/ModB")
        xlsx = _make_xlsx_with_sheets(tmp_path, [(sheet_a, None), (sheet_b, None)])

        project_cfg = _make_project_cfg(["/proj/ModA", "/proj/ModB"])
        db_path = tmp_path / "test.db"

        with patch("doors_excel.api.import_.stage_import") as mock_stage:
            mock_stage.return_value = ("sid-a", MagicMock())
            results = bulk_stage_imports(
                xlsx,
                project_cfg,
                db_path=db_path,
                doors_conn=MagicMock(),
                trim_whitespace=True,
            )

        assert len(results) == 2

    def test_bulk_stage_imports_uses_correct_sheet_for_prefix_modules(self, tmp_path: Path) -> None:
        """bulk_stage_imports must pass sheet_title to stage_import to bypass _pick_worksheet."""
        from doors_excel.api.import_ import bulk_stage_imports
        from doors_excel.infrastructure.excel.naming import make_sheet_name

        # Create two sheets whose module names share a prefix
        sheet_req = make_sheet_name("Req", "/proj/Req")
        sheet_req_safety = make_sheet_name("ReqSafety", "/proj/ReqSafety")
        xlsx = _make_xlsx_with_sheets(tmp_path, [(sheet_req, None), (sheet_req_safety, None)])

        project_cfg = _make_project_cfg(["/proj/Req", "/proj/ReqSafety"])
        db_path = tmp_path / "test.db"

        staged_titles: list[str | None] = []

        def capture_stage(excel_path, mod_cfg, *, db_path, doors_conn, trim_whitespace, sheet_title=None):
            staged_titles.append(sheet_title)
            return ("sid", MagicMock())

        with patch("doors_excel.api.import_.stage_import", side_effect=capture_stage):
            bulk_stage_imports(
                xlsx,
                project_cfg,
                db_path=db_path,
                doors_conn=MagicMock(),
                trim_whitespace=True,
            )

        assert sheet_req in staged_titles
        assert sheet_req_safety in staged_titles

    def test_bulk_stage_imports_skips_unmatched_sheets(self, tmp_path: Path) -> None:
        from doors_excel.api.import_ import bulk_stage_imports
        from doors_excel.infrastructure.excel.naming import make_sheet_name

        sheet_a = make_sheet_name("ModA", "/proj/ModA")
        xlsx = _make_xlsx_with_sheets(tmp_path, [(sheet_a, None), ("Unrelated", None)])

        project_cfg = _make_project_cfg(["/proj/ModA"])
        db_path = tmp_path / "test.db"

        with patch("doors_excel.api.import_.stage_import") as mock_stage:
            mock_stage.return_value = ("sid-a", MagicMock())
            results = bulk_stage_imports(
                xlsx,
                project_cfg,
                db_path=db_path,
                doors_conn=MagicMock(),
                trim_whitespace=True,
            )

        assert len(results) == 1
