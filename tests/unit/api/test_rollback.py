"""Unit tests for api/rollback.py — SQLite snapshot → Excel generation."""
from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
import pytest

from doors_excel.infrastructure.database.schema import apply_schema


@pytest.fixture
def conn() -> sqlite3.Connection:
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    apply_schema(c)
    c.execute(
        "INSERT INTO sessions (session_id, excel_path, doors_module, excel_sha256, module_version)"
        " VALUES (?,?,?,?,?)",
        ("sid1", "/tmp/data.xlsx", "/proj/mod", "aabbcc", "current"),
    )
    c.commit()
    return c


def _insert_snapshot(
    conn: sqlite3.Connection,
    session_id: str,
    object_id: int,
    attribute: str,
    value: str,
    rtf: str = "",
) -> None:
    conn.execute(
        "INSERT INTO rollback_snapshots (session_id, object_id, attribute, original_value, original_rtf)"
        " VALUES (?,?,?,?,?)",
        (session_id, object_id, attribute, value, rtf),
    )
    conn.commit()


class TestGenerateRollbackExcel:
    def test_returns_path_to_xlsx(self, conn: sqlite3.Connection, tmp_path: Path) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        _insert_snapshot(conn, "sid1", 1, "Object Text", "Hello")
        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_excel_contains_header_row(self, conn: sqlite3.Connection, tmp_path: Path) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        _insert_snapshot(conn, "sid1", 1, "Object Text", "Hello")
        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Absolute Number" in headers
        assert "Object Text" in headers

    def test_one_object_one_attribute_produces_data_row(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        _insert_snapshot(conn, "sid1", 42, "Object Text", "Hello world")
        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        row_vals = list(rows[0])
        assert 42 in row_vals
        assert "Hello world" in row_vals

    def test_multiple_attributes_pivoted_to_columns(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        _insert_snapshot(conn, "sid1", 1, "Object Text", "Title")
        _insert_snapshot(conn, "sid1", 1, "Short Name", "T")
        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Object Text" in headers
        assert "Short Name" in headers
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1

    def test_multiple_objects_produce_multiple_rows(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        for oid in [1, 2, 3]:
            _insert_snapshot(conn, "sid1", oid, "Object Text", f"Object {oid}")
        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3

    def test_empty_snapshots_produces_header_only(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        out = generate_rollback_excel("sid1", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 1

    def test_creates_output_directory_if_missing(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        _insert_snapshot(conn, "sid1", 1, "Object Text", "Hi")
        out = generate_rollback_excel("sid1", conn, tmp_path / "sub" / "rollback.xlsx")
        assert out.exists()

    def test_unknown_session_returns_empty_excel(
        self, conn: sqlite3.Connection, tmp_path: Path
    ) -> None:
        from doors_excel.api.rollback import generate_rollback_excel

        out = generate_rollback_excel("nosuchsession", conn, tmp_path / "rollback.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 1
