"""Tests for bulk_export — multi-module workbook (REQ-FUN-113)."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from doors_excel.core.validation.models import ColumnMapping, ModuleConfig


def _make_mod_cfg(path: str, col: str = "Object Text") -> ModuleConfig:
    return ModuleConfig(
        module_path=path,
        column_mappings=[
            ColumnMapping(column=col, attribute=col, attribute_type="Text"),
        ],
    )


def _raw_rows(module_path: str, object_id: int = 1) -> list[dict]:
    return [
        {
            "object_id": object_id,
            "level": 1,
            "parent_id": None,
            "has_ole": 0,
            "object_type": "OBJECT",
            "attribute": "Object Text",
            "value": f"Value from {module_path}",
            "rtf_value": "",
            "md_hash": None,
        }
    ]


class TestBulkExport:
    def test_bulk_export_creates_multisheet_workbook(self, tmp_path: Path) -> None:
        from doors_excel.api.export import bulk_export

        mod_a = _make_mod_cfg("/proj/ModA")
        mod_b = _make_mod_cfg("/proj/ModB")
        out = tmp_path / "bulk.xlsx"

        with patch("doors_excel.api.export.DoorsExporter") as MockExp, \
             patch("doors_excel.api.export._populate_session"):
            inst = MockExp.return_value
            inst.export_module.side_effect = [
                _raw_rows("/proj/ModA"),
                _raw_rows("/proj/ModB", object_id=2),
            ]
            result = bulk_export(
                [mod_a, mod_b],
                out,
                doors_conn=object(),
                session_manager=None,
            )

        wb = openpyxl.load_workbook(result)
        assert len(wb.sheetnames) == 2

    def test_bulk_export_returns_output_path(self, tmp_path: Path) -> None:
        from doors_excel.api.export import bulk_export

        mod_a = _make_mod_cfg("/proj/ModA")
        out = tmp_path / "bulk.xlsx"

        with patch("doors_excel.api.export.DoorsExporter") as MockExp, \
             patch("doors_excel.api.export._populate_session"):
            inst = MockExp.return_value
            inst.export_module.return_value = _raw_rows("/proj/ModA")
            result = bulk_export(
                [mod_a],
                out,
                doors_conn=object(),
                session_manager=None,
            )

        assert result == out
        assert result.exists()

    def test_bulk_export_each_sheet_has_data_rows(self, tmp_path: Path) -> None:
        from doors_excel.api.export import bulk_export

        mod_a = _make_mod_cfg("/proj/ModA")
        out = tmp_path / "bulk.xlsx"

        with patch("doors_excel.api.export.DoorsExporter") as MockExp, \
             patch("doors_excel.api.export._populate_session"):
            inst = MockExp.return_value
            inst.export_module.return_value = _raw_rows("/proj/ModA", object_id=42)
            result = bulk_export(
                [mod_a],
                out,
                doors_conn=object(),
                session_manager=None,
            )

        wb = openpyxl.load_workbook(result)
        ws = wb.worksheets[0]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        assert 42 in data_rows[0]

    def test_bulk_export_does_not_populate_sessions(self, tmp_path: Path) -> None:
        """bulk_export must not call _populate_session (sidecar overwrite limitation)."""
        from doors_excel.api.export import bulk_export
        from doors_excel.api.sessions import SessionManager

        with patch("doors_excel.api.export.DoorsExporter") as MockExp, \
             patch("doors_excel.api.export._populate_session") as mock_pop:
            MockExp.return_value.export_module.return_value = _raw_rows("/proj/ModA")
            bulk_export(
                [_make_mod_cfg("/proj/ModA")],
                tmp_path / "bulk.xlsx",
                doors_conn=object(),
                session_manager=SessionManager.__new__(SessionManager),  # dummy, must not be called
            )
        mock_pop.assert_not_called()

    def test_bulk_export_cli_flag(self, tmp_path: Path) -> None:
        """--bulk routes to bulk_export_api with project_cfg.modules and correct kwargs."""
        import json
        from typer.testing import CliRunner
        from doors_excel.cli.app import app

        cfg_data = {
            "modules": [
                {
                    "module_path": "/proj/ModA",
                    "column_mappings": [
                        {"column": "Object Text", "attribute": "Object Text", "attribute_type": "Text"}
                    ],
                },
            ]
        }
        cfg_file = tmp_path / "cfg.json"
        cfg_file.write_text(json.dumps(cfg_data))
        out = tmp_path / "bulk.xlsx"

        runner = CliRunner()
        with patch("doors_excel.cli.app.DoorsConnection") as MockConn, \
             patch("doors_excel.cli.app.bulk_export_api") as mock_bulk, \
             patch("doors_excel.cli.app.KeepAliveWatchdog"):
            MockConn.open.return_value.__enter__ = lambda s: MockConn.open.return_value
            MockConn.open.return_value.__exit__ = lambda *a: None
            mock_bulk.return_value = out
            result = runner.invoke(app, [
                "export", "--config", str(cfg_file), "--output", str(out), "--bulk",
            ])

        assert result.exit_code == 0
        mock_bulk.assert_called_once()
        call_args = mock_bulk.call_args
        # bulk_export_api is called with positional args: (modules, out_path, ...)
        modules_arg = call_args.args[0] if call_args.args else call_args.kwargs.get("module_configs", [])
        assert len(modules_arg) == 1
        assert modules_arg[0].module_path == "/proj/ModA"
        # Verify key kwargs are forwarded correctly
        kwargs = call_args.kwargs
        assert kwargs.get("baseline") == "current"
        assert kwargs.get("sheet_protection") is False
        assert kwargs.get("sheet_protection_password") is None
