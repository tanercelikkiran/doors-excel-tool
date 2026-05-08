"""Tests for api/sessions.py — SessionManager and helpers."""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from doors_excel.api.sessions import (
    SessionInfo,
    SessionManager,
    compute_file_sha256,
    session_file_path,
)
from doors_excel.common.constants import SESSION_FILE_NAME
from doors_excel.common.exceptions import SessionError
from doors_excel.infrastructure.database.schema import apply_schema


# ---------------------------------------------------------------------------
# compute_file_sha256
# ---------------------------------------------------------------------------

class TestComputeFileSha256:
    def test_returns_64_hex_chars(self, tmp_path: Path) -> None:
        f = tmp_path / "file.bin"
        f.write_bytes(b"hello world")
        h = compute_file_sha256(f)
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_deterministic(self, tmp_path: Path) -> None:
        f = tmp_path / "file.bin"
        f.write_bytes(b"deterministic")
        assert compute_file_sha256(f) == compute_file_sha256(f)

    def test_different_content_different_hash(self, tmp_path: Path) -> None:
        a = tmp_path / "a.bin"
        b = tmp_path / "b.bin"
        a.write_bytes(b"aaa")
        b.write_bytes(b"bbb")
        assert compute_file_sha256(a) != compute_file_sha256(b)


# ---------------------------------------------------------------------------
# session_file_path
# ---------------------------------------------------------------------------

class TestSessionFilePath:
    def test_returns_sidecar_next_to_xlsx(self, tmp_path: Path) -> None:
        p = tmp_path / "data.xlsx"
        sf = session_file_path(p)
        assert sf.parent == tmp_path
        assert sf.name == SESSION_FILE_NAME


# ---------------------------------------------------------------------------
# SessionInfo
# ---------------------------------------------------------------------------

class TestSessionInfo:
    def test_round_trip_dict(self, tmp_path: Path) -> None:
        info = SessionInfo(
            session_id="abc-123",
            excel_path=tmp_path / "f.xlsx",
            doors_module="/proj/mod",
            excel_sha256="deadbeef",
            module_version="current",
            db_path=tmp_path / "db.sqlite",
        )
        d = info.to_dict()
        restored = SessionInfo.from_dict(d)
        assert restored.session_id == info.session_id
        assert restored.excel_path == info.excel_path
        assert restored.doors_module == info.doors_module
        assert restored.excel_sha256 == info.excel_sha256
        assert restored.db_path == info.db_path

    def test_paths_are_path_objects(self, tmp_path: Path) -> None:
        info = SessionInfo.from_dict({
            "session_id": "x",
            "excel_path": str(tmp_path / "f.xlsx"),
            "doors_module": "/m",
            "excel_sha256": "abc",
            "module_version": "current",
            "db_path": str(tmp_path / "db.sqlite"),
        })
        assert isinstance(info.excel_path, Path)
        assert isinstance(info.db_path, Path)


# ---------------------------------------------------------------------------
# SessionManager.create
# ---------------------------------------------------------------------------

class TestSessionManagerCreate:
    def _make_xlsx(self, tmp_path: Path) -> Path:
        p = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        return p

    def test_returns_session_info(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            info = mgr.create(xlsx, "/proj/mod")
        assert info.session_id
        assert info.excel_path == xlsx
        assert len(info.excel_sha256) == 64

    def test_session_persisted_in_db(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            info = mgr.create(xlsx, "/proj/mod")
        # Verify row exists in DB
        conn = sqlite3.connect(str(db))
        apply_schema(conn)
        row = conn.execute(
            "SELECT * FROM sessions WHERE session_id = ?", (info.session_id,)
        ).fetchone()
        conn.close()
        assert row is not None

    def test_sidecar_json_created(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            info = mgr.create(xlsx, "/proj/mod")
        sf = session_file_path(xlsx)
        assert sf.exists()
        data = json.loads(sf.read_text())
        assert data["session_id"] == info.session_id

    def test_nonexistent_excel_raises_session_error(self, tmp_path: Path) -> None:
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            with pytest.raises(SessionError, match="not found"):
                mgr.create(tmp_path / "missing.xlsx", "/proj/mod")

    def test_unique_session_ids(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            a = mgr.create(xlsx, "/proj/mod")
            b = mgr.create(xlsx, "/proj/mod")
        assert a.session_id != b.session_id


# ---------------------------------------------------------------------------
# SessionManager.resume
# ---------------------------------------------------------------------------

class TestSessionManagerResume:
    def _make_xlsx(self, tmp_path: Path) -> Path:
        p = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        return p

    def test_resume_returns_matching_info(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            created = mgr.create(xlsx, "/proj/mod")
        with SessionManager(db) as mgr:
            resumed = mgr.resume(session_file_path(xlsx))
        assert resumed.session_id == created.session_id

    def test_resume_fails_on_missing_session_file(self, tmp_path: Path) -> None:
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            with pytest.raises(SessionError, match="Cannot read"):
                mgr.resume(tmp_path / "nonexistent.json")

    def test_resume_fails_on_modified_excel(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            mgr.create(xlsx, "/proj/mod")
        # Tamper the file
        xlsx.write_bytes(b"corrupted content that changes the hash")
        with SessionManager(db) as mgr:
            with pytest.raises(SessionError, match="modified"):
                mgr.resume(session_file_path(xlsx))

    def test_resume_fails_if_excel_missing(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            mgr.create(xlsx, "/proj/mod")
        xlsx.unlink()
        with SessionManager(db) as mgr:
            with pytest.raises(SessionError, match="no longer exists"):
                mgr.resume(session_file_path(tmp_path / "data.xlsx"))


# ---------------------------------------------------------------------------
# SessionManager.finish / fail
# ---------------------------------------------------------------------------

class TestSessionManagerStatus:
    def _make_xlsx(self, tmp_path: Path) -> Path:
        p = tmp_path / "data.xlsx"
        openpyxl.Workbook().save(p)
        return p

    def test_finish_sets_completed(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            info = mgr.create(xlsx, "/proj/mod")
            mgr.finish(info.session_id)
            row = mgr.conn.execute(
                "SELECT status FROM sessions WHERE session_id = ?", (info.session_id,)
            ).fetchone()
        assert row[0] == "completed"

    def test_fail_sets_failed(self, tmp_path: Path) -> None:
        xlsx = self._make_xlsx(tmp_path)
        db = tmp_path / "test.db"
        with SessionManager(db) as mgr:
            info = mgr.create(xlsx, "/proj/mod")
            mgr.fail(info.session_id)
            row = mgr.conn.execute(
                "SELECT status FROM sessions WHERE session_id = ?", (info.session_id,)
            ).fetchone()
        assert row[0] == "failed"
