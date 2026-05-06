"""Unit + round-trip tests for Excel sheet-protection helpers (REQ-FUN-108.7)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from doors_excel.infrastructure.excel.protection import (
    apply_sheet_protection,
    lock_cell,
    lock_range,
    unlock_cell,
)
from doors_excel.infrastructure.excel.writer import add_module_sheet, create_workbook, save_workbook


# ---------------------------------------------------------------------------
# lock_cell / unlock_cell
# ---------------------------------------------------------------------------

class TestLockUnlockCell:
    def test_lock_cell_sets_protection(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws["A1"] = "value"
        lock_cell(ws, "A1")
        assert ws["A1"].protection.locked is True

    def test_unlock_cell_clears_protection(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws["A1"] = "value"
        lock_cell(ws, "A1")
        unlock_cell(ws, "A1")
        assert ws["A1"].protection.locked is False

    def test_lock_cell_accepts_row_col(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws.cell(row=2, column=3).value = "x"
        lock_cell(ws, row=2, col=3)
        assert ws.cell(row=2, column=3).protection.locked is True

    def test_unlock_cell_accepts_row_col(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws.cell(row=1, column=1).value = "x"
        lock_cell(ws, row=1, col=1)
        unlock_cell(ws, row=1, col=1)
        assert ws.cell(row=1, column=1).protection.locked is False

    def test_lock_requires_cell_ref_or_row_col(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        with pytest.raises((TypeError, ValueError)):
            lock_cell(ws)  # type: ignore[call-arg]


# ---------------------------------------------------------------------------
# lock_range
# ---------------------------------------------------------------------------

class TestLockRange:
    def test_locks_all_cells_in_range(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c).value = f"{r},{c}"
        lock_range(ws, min_row=1, max_row=3, min_col=1, max_col=3)
        for r in range(1, 4):
            for c in range(1, 4):
                assert ws.cell(row=r, column=c).protection.locked is True

    def test_lock_range_single_row(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws["A1"] = "h1"
        ws["B1"] = "h2"
        lock_range(ws, min_row=1, max_row=1, min_col=1, max_col=2)
        assert ws["A1"].protection.locked is True
        assert ws["B1"].protection.locked is True


# ---------------------------------------------------------------------------
# apply_sheet_protection
# ---------------------------------------------------------------------------

class TestApplySheetProtection:
    def test_protection_enabled(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        assert ws.protection.sheet is True

    def test_insert_rows_disabled(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        # insertRows protection flag: True means the action is BLOCKED
        assert ws.protection.insertRows is True

    def test_delete_rows_disabled(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        assert ws.protection.deleteRows is True

    def test_select_locked_cells_allowed(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        # selectLockedCells=False means selecting locked cells IS allowed
        assert ws.protection.selectLockedCells is False

    def test_select_unlocked_cells_allowed(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        assert ws.protection.selectUnlockedCells is False

    def test_password_set_when_provided(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws, password="secret")
        # openpyxl stores a hashed password, not plaintext
        assert ws.protection.password is not None
        assert ws.protection.password != ""

    def test_no_password_by_default(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws)
        # Without a password, the hash field should be empty/None
        assert not ws.protection.password

    def test_round_trip_protection_survives(self, tmp_path: Path) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        ws["A1"] = "data"
        lock_cell(ws, "A1")
        apply_sheet_protection(ws)
        out = tmp_path / "protected.xlsx"
        save_workbook(wb, out)

        wb2 = openpyxl.load_workbook(out, data_only=True)
        ws2 = wb2.worksheets[0]
        assert ws2.protection.sheet is True
        assert ws2["A1"].protection.locked is True

    def test_disabled_protection(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "S", "/s")
        apply_sheet_protection(ws, enabled=False)
        assert ws.protection.sheet is False
