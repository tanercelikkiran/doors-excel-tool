"""Unit + round-trip tests for Excel metadata helpers (REQ-FUN-113.2, REQ-FUN-108.1, REQ-FUN-117)."""
from __future__ import annotations

import hashlib
import json
import tempfile
from pathlib import Path

import openpyxl
import pytest

from doors_excel.infrastructure.excel.metadata import (
    DOORS_MODULE_PATH_DEFINED_NAME,
    compute_integrity_hash,
    get_custom_property,
    get_module_path,
    set_custom_property,
    set_module_path,
    verify_integrity_hash,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_wb_with_sheet(sheet_name: str = "Sheet1") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb


def _round_trip(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """Save to a temp file, reload, and return the reopened workbook."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    wb.save(tmp)
    return openpyxl.load_workbook(tmp, data_only=True)


# ---------------------------------------------------------------------------
# DOORS_MODULE_PATH_DEFINED_NAME constant
# ---------------------------------------------------------------------------

class TestConstant:
    def test_constant_value(self) -> None:
        assert DOORS_MODULE_PATH_DEFINED_NAME == "_DOORS_Module_Path"


# ---------------------------------------------------------------------------
# set_module_path / get_module_path  (REQ-FUN-113.2)
# ---------------------------------------------------------------------------

class TestModulePathDefinedName:
    def test_set_and_get_in_memory(self) -> None:
        wb = _make_wb_with_sheet("SRS")
        ws = wb["SRS"]
        set_module_path(wb, ws, "/Project/Folder/SRS")
        assert get_module_path(wb, ws) == "/Project/Folder/SRS"

    def test_missing_returns_none(self) -> None:
        wb = _make_wb_with_sheet("Sheet1")
        ws = wb.active
        assert get_module_path(wb, ws) is None

    def test_round_trip_file(self, tmp_path: Path) -> None:
        wb = _make_wb_with_sheet("TestSheet")
        ws = wb["TestSheet"]
        set_module_path(wb, ws, "/Project/TestModule")
        out = tmp_path / "test.xlsx"
        wb.save(out)
        wb2 = openpyxl.load_workbook(out, data_only=True)
        ws2 = wb2["TestSheet"]
        assert get_module_path(wb2, ws2) == "/Project/TestModule"

    def test_overwrite_updates_value(self) -> None:
        wb = _make_wb_with_sheet("S1")
        ws = wb.active
        set_module_path(wb, ws, "/old/path")
        set_module_path(wb, ws, "/new/path")
        assert get_module_path(wb, ws) == "/new/path"

    def test_two_sheets_independent(self) -> None:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "A"
        ws2 = wb.create_sheet("B")
        set_module_path(wb, ws1, "/mod/A")
        set_module_path(wb, ws2, "/mod/B")
        assert get_module_path(wb, ws1) == "/mod/A"
        assert get_module_path(wb, ws2) == "/mod/B"


# ---------------------------------------------------------------------------
# set_custom_property / get_custom_property  (REQ-FUN-108.1)
# ---------------------------------------------------------------------------

class TestCustomProperties:
    def test_set_and_get_string(self, tmp_path: Path) -> None:
        wb = _make_wb_with_sheet()
        set_custom_property(wb, "MarkerMapping", '{"TABLE_START": "TABLE_START"}')
        out = tmp_path / "cp.xlsx"
        wb.save(out)
        wb2 = openpyxl.load_workbook(out, data_only=True)
        assert get_custom_property(wb2, "MarkerMapping") == '{"TABLE_START": "TABLE_START"}'

    def test_missing_returns_none(self) -> None:
        wb = _make_wb_with_sheet()
        assert get_custom_property(wb, "NonExistent") is None

    def test_overwrite_custom_property(self, tmp_path: Path) -> None:
        wb = _make_wb_with_sheet()
        set_custom_property(wb, "Key", "v1")
        set_custom_property(wb, "Key", "v2")
        out = tmp_path / "ow.xlsx"
        wb.save(out)
        wb2 = openpyxl.load_workbook(out, data_only=True)
        assert get_custom_property(wb2, "Key") == "v2"

    def test_multiple_properties(self, tmp_path: Path) -> None:
        wb = _make_wb_with_sheet()
        set_custom_property(wb, "Alpha", "aaa")
        set_custom_property(wb, "Beta", "bbb")
        out = tmp_path / "multi.xlsx"
        wb.save(out)
        wb2 = openpyxl.load_workbook(out, data_only=True)
        assert get_custom_property(wb2, "Alpha") == "aaa"
        assert get_custom_property(wb2, "Beta") == "bbb"


# ---------------------------------------------------------------------------
# compute_integrity_hash / verify_integrity_hash  (REQ-FUN-117)
# ---------------------------------------------------------------------------

class TestIntegrityHash:
    def _fields(self) -> dict[str, str]:
        return {
            "Source Baseline": "1.2",
            "_DOORS_Module_Path": "/Project/SRS",
            "AbsNum_1": "42",
        }

    def test_returns_64_char_hex(self) -> None:
        h = compute_integrity_hash(self._fields())
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_deterministic(self) -> None:
        assert compute_integrity_hash(self._fields()) == compute_integrity_hash(self._fields())

    def test_different_fields_different_hash(self) -> None:
        fields2 = {**self._fields(), "AbsNum_1": "99"}
        assert compute_integrity_hash(self._fields()) != compute_integrity_hash(fields2)

    def test_key_order_independent(self) -> None:
        # Sorted keys: same hash regardless of insertion order
        fields_a = {"b": "2", "a": "1"}
        fields_b = {"a": "1", "b": "2"}
        assert compute_integrity_hash(fields_a) == compute_integrity_hash(fields_b)

    def test_verify_returns_true_for_matching_hash(self) -> None:
        fields = self._fields()
        h = compute_integrity_hash(fields)
        assert verify_integrity_hash(fields, h) is True

    def test_verify_returns_false_for_tampered_fields(self) -> None:
        fields = self._fields()
        h = compute_integrity_hash(fields)
        fields["AbsNum_1"] = "99"
        assert verify_integrity_hash(fields, h) is False

    def test_verify_returns_false_for_wrong_hash(self) -> None:
        assert verify_integrity_hash(self._fields(), "0" * 64) is False
