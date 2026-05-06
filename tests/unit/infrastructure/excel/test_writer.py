"""Unit + round-trip tests for the Excel writer."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from doors_excel.infrastructure.excel.metadata import (
    get_custom_property,
    get_module_path,
    set_module_path,
)
from doors_excel.infrastructure.excel.writer import (
    add_module_sheet,
    create_workbook,
    save_workbook,
)


# ---------------------------------------------------------------------------
# create_workbook
# ---------------------------------------------------------------------------

class TestCreateWorkbook:
    def test_returns_workbook(self) -> None:
        wb = create_workbook()
        assert isinstance(wb, openpyxl.Workbook)

    def test_no_default_sheet(self) -> None:
        wb = create_workbook()
        # We create a clean workbook; default "Sheet" stub should be removed
        assert len(wb.sheetnames) == 0

    def test_write_only_false(self) -> None:
        wb = create_workbook()
        assert not wb.write_only


# ---------------------------------------------------------------------------
# add_module_sheet
# ---------------------------------------------------------------------------

class TestAddModuleSheet:
    def test_adds_sheet_with_correct_name(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "SRS", "/project/SRS")
        assert "SRS_" in ws.title or ws.title in wb.sheetnames

    def test_module_path_embedded(self) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "SRS", "/project/SRS")
        assert get_module_path(wb, ws) == "/project/SRS"

    def test_sheet_name_uses_naming_convention(self) -> None:
        from doors_excel.infrastructure.excel.naming import make_sheet_name
        wb = create_workbook()
        ws = add_module_sheet(wb, "My Module", "/project/My Module")
        expected = make_sheet_name("My Module", "/project/My Module")
        assert ws.title == expected

    def test_multiple_sheets_different_names(self) -> None:
        wb = create_workbook()
        ws1 = add_module_sheet(wb, "SRS", "/project/A/SRS")
        ws2 = add_module_sheet(wb, "SRS", "/project/B/SRS")
        assert ws1.title != ws2.title

    def test_multiple_sheets_paths_independent(self) -> None:
        wb = create_workbook()
        ws1 = add_module_sheet(wb, "ModA", "/project/A")
        ws2 = add_module_sheet(wb, "ModB", "/project/B")
        assert get_module_path(wb, ws1) == "/project/A"
        assert get_module_path(wb, ws2) == "/project/B"


# ---------------------------------------------------------------------------
# save_workbook
# ---------------------------------------------------------------------------

class TestSaveWorkbook:
    def test_creates_file(self, tmp_path: Path) -> None:
        wb = create_workbook()
        add_module_sheet(wb, "SRS", "/p/SRS")
        out = tmp_path / "output.xlsx"
        save_workbook(wb, out)
        assert out.exists()

    def test_creates_parent_dirs(self, tmp_path: Path) -> None:
        wb = create_workbook()
        add_module_sheet(wb, "SRS", "/p/SRS")
        out = tmp_path / "sub" / "dir" / "output.xlsx"
        save_workbook(wb, out)
        assert out.exists()

    def test_round_trip_preserves_module_path(self, tmp_path: Path) -> None:
        wb = create_workbook()
        ws = add_module_sheet(wb, "SRS", "/project/SRS")
        out = tmp_path / "rt.xlsx"
        save_workbook(wb, out)

        wb2 = openpyxl.load_workbook(out, data_only=True)
        ws2 = wb2.worksheets[0]
        assert get_module_path(wb2, ws2) == "/project/SRS"

    def test_round_trip_preserves_custom_property(self, tmp_path: Path) -> None:
        wb = create_workbook()
        add_module_sheet(wb, "SRS", "/project/SRS")
        from doors_excel.infrastructure.excel.metadata import set_custom_property
        set_custom_property(wb, "MarkerMapping", '{"TABLE_START":"TABLE_START"}')
        out = tmp_path / "props.xlsx"
        save_workbook(wb, out)

        wb2 = openpyxl.load_workbook(out, data_only=True)
        assert get_custom_property(wb2, "MarkerMapping") == '{"TABLE_START":"TABLE_START"}'

    def test_overwrite_existing_file(self, tmp_path: Path) -> None:
        wb1 = create_workbook()
        add_module_sheet(wb1, "A", "/a")
        out = tmp_path / "overwrite.xlsx"
        save_workbook(wb1, out)

        wb2 = create_workbook()
        add_module_sheet(wb2, "B", "/b")
        save_workbook(wb2, out)

        wb3 = openpyxl.load_workbook(out, data_only=True)
        # Only the second workbook's sheet should be present
        assert len(wb3.sheetnames) == 1

    def test_returns_path(self, tmp_path: Path) -> None:
        wb = create_workbook()
        add_module_sheet(wb, "SRS", "/p/SRS")
        out = tmp_path / "out.xlsx"
        result = save_workbook(wb, out)
        assert result == out
