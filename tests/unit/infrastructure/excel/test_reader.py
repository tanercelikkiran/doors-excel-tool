"""Unit + round-trip tests for the Excel reader (REQ-FUN-215.1, C-7)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from doors_excel.infrastructure.excel.metadata import set_custom_property, set_module_path
from doors_excel.infrastructure.excel.naming import make_sheet_name
from doors_excel.infrastructure.excel.reader import (
    FormulaPolicy,
    open_workbook,
    resolve_sheet_module,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def simple_xlsx(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    out = tmp_path / "simple.xlsx"
    wb.save(out)
    return out


@pytest.fixture()
def formula_xlsx(tmp_path: Path) -> Path:
    """Workbook that contains a formula cell (cached value = 42)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 42  # openpyxl writes this as a plain value, simulating a cached result
    ws["B1"] = "=A1*2"
    out = tmp_path / "formula.xlsx"
    wb.save(out)
    return out


@pytest.fixture()
def module_path_xlsx(tmp_path: Path) -> Path:
    """Workbook with _DOORS_Module_Path metadata on one sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SRS"
    set_module_path(wb, ws, "/Project/SRS")
    out = tmp_path / "with_path.xlsx"
    wb.save(out)
    return out


@pytest.fixture()
def crc_sheet_xlsx(tmp_path: Path) -> Path:
    """Workbook whose sheet name follows the CRC32 convention."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = make_sheet_name("SRS", "/project/A/SRS")
    out = tmp_path / "crc.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# open_workbook
# ---------------------------------------------------------------------------

class TestOpenWorkbook:
    def test_returns_workbook(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        assert isinstance(wb, openpyxl.Workbook)

    def test_data_only_by_default(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        # data_only=True is the C-7 requirement; verify via the formula fixture
        # (openpyxl sets wb.data_only when loaded that way)
        assert wb.data_only is True

    def test_can_read_cell_value(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        assert wb.active["A1"].value == "hello"

    def test_missing_file_raises_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            open_workbook(tmp_path / "nonexistent.xlsx")

    def test_formula_policy_data_only_default(self, formula_xlsx: Path) -> None:
        wb = open_workbook(formula_xlsx, formula_policy=FormulaPolicy.DATA_ONLY)
        ws = wb.active
        # B1 was a formula; data_only=True returns the cached value (None if never evaluated)
        # We just check it doesn't return the raw formula string
        val = ws["B1"].value
        assert val != "=A1*2"

    def test_formula_policy_keep_formulas(self, formula_xlsx: Path) -> None:
        wb = open_workbook(formula_xlsx, formula_policy=FormulaPolicy.KEEP_FORMULAS)
        assert wb.data_only is False


# ---------------------------------------------------------------------------
# resolve_sheet_module  (REQ-FUN-215.1)
# ---------------------------------------------------------------------------

class TestResolveSheetModule:
    """4-tier resolution: metadata → config mapping → CRC32 → literal name."""

    # ---- Tier 1: embedded metadata ------------------------------------------

    def test_tier1_metadata_match(self, module_path_xlsx: Path) -> None:
        wb = open_workbook(module_path_xlsx)
        ws = wb["SRS"]
        result = resolve_sheet_module(wb, ws)
        assert result == "/Project/SRS"

    # ---- Tier 2: config mapping ---------------------------------------------

    def test_tier2_config_mapping(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        ws = wb.active
        mapping = {"Sheet1": "/Project/FromConfig"}
        result = resolve_sheet_module(wb, ws, workbook_mapping=mapping)
        assert result == "/Project/FromConfig"

    def test_tier2_not_used_when_tier1_present(self, module_path_xlsx: Path) -> None:
        wb = open_workbook(module_path_xlsx)
        ws = wb["SRS"]
        mapping = {"SRS": "/Project/FromConfig"}
        result = resolve_sheet_module(wb, ws, workbook_mapping=mapping)
        # Tier 1 wins
        assert result == "/Project/SRS"

    # ---- Tier 3: CRC32 match ------------------------------------------------

    def test_tier3_crc_match(self, crc_sheet_xlsx: Path) -> None:
        wb = open_workbook(crc_sheet_xlsx)
        ws = wb.active
        known_paths = ["/project/A/SRS", "/project/B/Other"]
        result = resolve_sheet_module(wb, ws, known_paths=known_paths)
        assert result == "/project/A/SRS"

    def test_tier3_no_match_if_no_known_paths(self, crc_sheet_xlsx: Path) -> None:
        wb = open_workbook(crc_sheet_xlsx)
        ws = wb.active
        result = resolve_sheet_module(wb, ws)
        assert result is None

    # ---- Tier 4: literal name -----------------------------------------------

    def test_tier4_literal_name_match(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        ws = wb.active  # title = "Sheet1"
        known_modules = {"sheet1": "/Project/ByName"}
        result = resolve_sheet_module(wb, ws, literal_modules=known_modules)
        assert result == "/Project/ByName"

    def test_tier4_case_insensitive(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        ws = wb.active
        known_modules = {"SHEET1": "/Project/ByName"}
        result = resolve_sheet_module(wb, ws, literal_modules=known_modules)
        assert result == "/Project/ByName"

    # ---- No match -----------------------------------------------------------

    def test_no_match_returns_none(self, simple_xlsx: Path) -> None:
        wb = open_workbook(simple_xlsx)
        ws = wb.active
        result = resolve_sheet_module(wb, ws)
        assert result is None
