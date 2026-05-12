"""Tests for feedback column writer (REQ-SAF-402)."""
from __future__ import annotations

import openpyxl


class TestWriteValidationFeedback:
    def _make_ws(self, data: list[list]):
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        return wb, ws

    def test_feedback_column_added(self):
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", "Object Text"],
            [1, "hello"],
            [2, "world"],
        ])
        # row_number=2 → first data row (Excel row 2)
        write_validation_feedback(ws, {2: ["Missing required field"]})
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert FEEDBACK_COLUMN in headers

    def test_feedback_written_to_correct_row(self):
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", "Object Text"],
            [1, "hello"],
            [2, "world"],
        ])
        # row_number=2 → Excel row 2 (first data row)
        # row_number=3 → Excel row 3 (second data row)
        write_validation_feedback(ws, {2: ["Bad value"], 3: ["Missing field"]})
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        assert "Bad value" in ws.cell(2, fb_col).value
        assert "Missing field" in ws.cell(3, fb_col).value

    def test_feedback_column_cleared_on_rewrite(self):
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", "Object Text"],
            [1, "hello"],
        ])
        write_validation_feedback(ws, {2: ["Error A"]})
        write_validation_feedback(ws, {})  # clear
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        # data cell should now be empty
        assert ws.cell(2, fb_col).value is None

    def test_no_errors_leaves_cells_empty(self):
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", "Object Text"],
            [1, "hello"],
            [2, "world"],
        ])
        write_validation_feedback(ws, {})
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        assert ws.cell(2, fb_col).value is None
        assert ws.cell(3, fb_col).value is None

    def test_sheet_level_errors_are_skipped(self):
        """row_number=0 (MISSING_COLUMN) must not write to any data row."""
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", "Object Text"],
            [1, "hello"],
        ])
        write_validation_feedback(ws, {0: ["Required column 'Foo' is absent."]})
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        # header row should have the column name but no data in row 2
        assert ws.cell(1, fb_col).value == FEEDBACK_COLUMN
        assert ws.cell(2, fb_col).value is None

    def test_existing_feedback_column_reused(self):
        """If the column already exists it should be reused, not appended."""
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([
            ["Absolute Number", FEEDBACK_COLUMN, "Object Text"],
            [1, "old error", "hello"],
        ])
        write_validation_feedback(ws, {2: ["new error"]})
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # Should still be 3 columns, not 4
        assert len(headers) == 3
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        assert ws.cell(2, fb_col).value == "new error"

    def test_feedback_roundtrip(self, tmp_path):
        """Feedback column survives workbook save and reload."""
        from doors_excel.infrastructure.excel.feedback import write_validation_feedback, FEEDBACK_COLUMN
        wb, ws = self._make_ws([["Absolute Number", "Object Text"], [1, "hello"]])
        write_validation_feedback(ws, {2: ["Round-trip error"]})
        path = tmp_path / "out.xlsx"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
        fb_col = headers.index(FEEDBACK_COLUMN) + 1
        assert ws2.cell(2, fb_col).value == "Round-trip error"
