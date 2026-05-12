"""Write _DOORS_Validation_Feedback column back to Excel (REQ-SAF-402)."""
from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

FEEDBACK_COLUMN = "_DOORS_Validation_Feedback"


def write_validation_feedback(
    ws: "Worksheet",
    row_errors: dict[int, list[str]],
    *,
    column_name: str = FEEDBACK_COLUMN,
) -> None:
    """Upsert feedback column in *ws*; clear previous data, write new messages.

    *row_errors* maps Excel row number → list of error strings.
    Row number 2 corresponds to the first data row (row 1 is the header).
    Row number 0 is used for sheet-level errors (e.g. MISSING_COLUMN); these
    are skipped because they do not correspond to any data row.
    """
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if column_name in headers:
        fb_col = headers.index(column_name) + 1
    else:
        fb_col = ws.max_column + 1
        ws.cell(1, fb_col).value = column_name

    # Clear existing data in feedback column
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, fb_col).value = None

    # Write new messages; row_number=0 are sheet-level errors (no data row)
    for excel_row, messages in row_errors.items():
        if excel_row < 2:
            continue  # skip sheet-level MISSING_COLUMN errors
        ws.cell(excel_row, fb_col).value = "; ".join(messages)
