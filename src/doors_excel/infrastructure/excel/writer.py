"""Excel workbook writer with metadata embedding."""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from doors_excel.infrastructure.excel.metadata import set_module_path
from doors_excel.infrastructure.excel.naming import make_sheet_name

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


def create_workbook() -> "Workbook":
    """Create a fresh openpyxl Workbook with no default sheets."""
    wb = openpyxl.Workbook()
    # openpyxl always creates an initial "Sheet" — remove it so callers
    # start with a completely empty workbook.
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


def add_module_sheet(
    wb: "Workbook",
    module_name: str,
    module_path: str,
) -> "Worksheet":
    """Add a new worksheet for *module_name* / *module_path* and embed metadata.

    The sheet title follows the ``make_sheet_name`` convention (REQ-FUN-113.1).
    The module path is stored as a Custom Property (REQ-FUN-113.2) so it
    survives save/load round-trips.
    """
    title = make_sheet_name(module_name, module_path)
    ws = wb.create_sheet(title=title)
    set_module_path(wb, ws, module_path)
    return ws


def save_workbook(wb: "Workbook", path: Path | str) -> Path:
    """Save *wb* to *path*, creating intermediate directories as needed.

    Returns the resolved ``Path`` to the saved file.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
