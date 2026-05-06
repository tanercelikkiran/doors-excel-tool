"""Cell-locking and sheet-protection helpers (REQ-FUN-108.7)."""
from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Protection

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def lock_cell(
    ws: "Worksheet",
    cell_ref: str | None = None,
    *,
    row: int | None = None,
    col: int | None = None,
) -> None:
    """Lock a single cell so it cannot be edited when sheet protection is active.

    Pass either *cell_ref* (e.g. ``"A1"``) **or** both *row* and *col*.
    """
    cell = _get_cell(ws, cell_ref, row, col)
    cell.protection = Protection(locked=True)


def unlock_cell(
    ws: "Worksheet",
    cell_ref: str | None = None,
    *,
    row: int | None = None,
    col: int | None = None,
) -> None:
    """Unlock a single cell so it remains editable when sheet protection is active."""
    cell = _get_cell(ws, cell_ref, row, col)
    cell.protection = Protection(locked=False)


def lock_range(
    ws: "Worksheet",
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """Lock all cells in the specified rectangular range."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).protection = Protection(locked=True)


def apply_sheet_protection(
    ws: "Worksheet",
    *,
    password: str | None = None,
    enabled: bool = True,
) -> None:
    """Apply (or remove) Excel sheet protection on *ws* (REQ-FUN-108.7).

    When *enabled* is ``True`` (the default):

    * Sheet is protected — users cannot modify locked cells.
    * Inserting and deleting rows are blocked.
    * Selecting both locked *and* unlocked cells is still allowed.

    *password* is optional; if supplied it is hashed and embedded so Excel
    requires it before the sheet can be unprotected.
    """
    p = ws.protection
    p.sheet = enabled

    if not enabled:
        return

    # Block structural mutations
    p.insertRows = True
    p.deleteRows = True

    # Allow selecting any cell (locked or not) — REQ-FUN-108.7
    p.selectLockedCells = False
    p.selectUnlockedCells = False

    if password is not None:
        p.set_password(password)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _get_cell(
    ws: "Worksheet",
    cell_ref: str | None,
    row: int | None,
    col: int | None,
):  # type: ignore[return]
    if cell_ref is not None:
        return ws[cell_ref]
    if row is not None and col is not None:
        return ws.cell(row=row, column=col)
    raise ValueError("Provide either cell_ref or both row and col")
